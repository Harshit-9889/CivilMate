from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.styles import Alignment
from openpyxl.styles import Border
from openpyxl.styles import Side
from datetime import datetime

thin_border = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin")
)

def export_concrete_report(
    project_name,
    safe_project_name,
    client_name,
    estimate_id,
    cement,
    sand,
    aggregate,
    cement_rate,
    sand_rate,
    aggregate_rate,
    costs
):
    today = datetime.now()
    current_date = today.strftime(
    "%d-%b-%Y"
    )

    wb = Workbook()

    ws = wb.active

    ws.title = "Concrete Estimate"

    ws.merge_cells("A1:E1")

    ws["A1"] = "CONCRETE ESTIMATION REPORT"
    ws["A1"].font = Font(
    bold=True,
    size=16
    )
    ws["A1"].alignment = Alignment(
    horizontal="center"
    )

    ws["A3"] = "Project Name"
    ws["B3"] = project_name

    ws["A4"] = "Client Name"
    ws["B4"] = client_name

    ws["A5"] = "Estimate ID"
    ws["B5"] = estimate_id

    ws["A6"] = "Date"
    ws["B6"] = current_date

    ws["A8"] = "Material"
    ws["B8"] = "Quantity"
    ws["C8"] = "Unit"
    ws["D8"] = "Rate"
    ws["E8"] = "Amount"

    for cell in ["A8", "B8", "C8", "D8", "E8"]:
        ws[cell].font = Font(bold=True)
    for cell in ["A8", "B8", "C8", "D8", "E8"]:
        ws[cell].alignment = Alignment(
        horizontal="center"
    )

    ws["A9"] = "Cement"
    ws["B9"] = cement
    ws["C9"] = "Bags"
    ws["D9"] = cement_rate
    ws["E9"] = costs["cement_cost"]

    ws["A10"] = "Sand"
    ws["B10"] = sand
    ws["C10"] = "m³"
    ws["D10"] = sand_rate
    ws["E10"] = costs["sand_cost"]

    ws["A11"] = "Aggregate"
    ws["B11"] = aggregate
    ws["C11"] = "m³"
    ws["D11"] = aggregate_rate
    ws["E11"] = costs["aggregate_cost"]

    ws["A12"] = "TOTAL COST"
    ws["E12"] =costs["total_cost"]

    for cell in ["A12", "B12", "C12", "D12", "E12"]:
        ws[cell].font = Font(
        bold=True
    )
    
    for cell in ["A12", "B12", "C12", "D12", "E12"]:
        ws[cell].alignment = Alignment(
        horizontal="center"
    )

    for row in range(9,12):
        for col in ["A", "B", "C", "D", "E"]:
            ws[f"{col}{row}"].alignment = Alignment(
            horizontal="center"
        )
    
    for row in range(8, 13):
        for col in ["A", "B", "C", "D", "E"]:
         ws[f"{col}{row}"].border = thin_border

    ws.column_dimensions["A"].width = 20
    ws.column_dimensions["B"].width = 15
    ws.column_dimensions["C"].width = 15
    ws.column_dimensions["D"].width = 15
    ws.column_dimensions["E"].width = 15
 
    filename = (
    f"{safe_project_name}_{estimate_id}.xlsx"
    )

    wb.save(
    f"reports/Excel/{filename}"
    )

    print(
        "Excel Report Saved Successfully!"
    )
