from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.styles import Alignment
from openpyxl.styles import Border
from openpyxl.styles import Side
from datetime import datetime

thin_border = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin")
)
def export_steelwork_report (
    project_name,
    safe_project_name,
    client_name,
    estimate_id,
    steel_weight,
    bar_type,
    steel_rate,
    steel_cost     
):
    today = datetime.now()
    current_date = today.strftime(
    "%d-%b-%Y"
    )

    wb = Workbook()

    ws = wb.active

    ws.title = "Steel Estimate"

    ws.merge_cells("A1:E1")

    ws["A1"] = "STEEL ESTIMATION REPORT"
    ws["A1"].font = Font(
    bold=True,
    size=16
    )
    ws["A1"].alignment = Alignment(
    horizontal="center"
    )

    ws["A3"] = "Project Name"
    ws["B3"] = project_name

    ws["A4"] = "Client Name"
    ws["B4"] = client_name

    ws["A5"] = "Estimate ID"
    ws["B5"] = estimate_id

    ws["A6"] = "Date"
    ws["B6"] = current_date

    ws["A8"] = "Description"
    ws["B8"] = "Quantity"
    ws["C8"] = "Unit"
    ws["D8"] = "Bar Type"
    ws["E8"] = "Rate"
    ws["F8"] = "Amount"

    for cell in ["A8", "B8", "C8", "D8", "E8","F8"]:
        ws[cell].font = Font(bold=True)
    for cell in ["A8", "B8", "C8", "D8", "E8","F8"]:
        ws[cell].alignment = Alignment(
        horizontal="center"
        )

    ws["A9"] = "Steel Reinforcement"
    ws["B9"] = steel_weight
    ws["C9"] = "kg"
    ws["D9"] = bar_type
    ws["E9"] = steel_rate
    ws["F9"] = steel_cost

    ws["A10"] = "TOTAL COST"
    ws["F10"] = steel_cost

    for cell in ["A10", "B10", "C10", "D10", "E10", "F10"]:
        ws[cell].font = Font(
        bold=True
    )
    
    for cell in ["A10", "B10", "C10", "D10", "E10", "F10"]:
        ws[cell].alignment = Alignment(
        horizontal="center"
    )
    for cell in [ "A9", "B9", "C9", "D9", "E9", "F9"]:
            ws[cell].alignment = Alignment(
            horizontal="center"
        )
    
    for row in range(8, 11):
        for col in ["A", "B", "C", "D", "E" , "F"]:
         ws[f"{col}{row}"].border = thin_border

    ws.column_dimensions["A"].width = 20
    ws.column_dimensions["B"].width = 15
    ws.column_dimensions["C"].width = 15
    ws.column_dimensions["D"].width = 15
    ws.column_dimensions["E"].width = 15

    filename = (
    f"{safe_project_name}_{estimate_id}.xlsx"
    )

    wb.save(
    f"reports/Excel/{filename}"
    )

    print(
        "Excel Report Saved Successfully!"
    )
    
